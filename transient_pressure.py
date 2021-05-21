import numpy as np
import openpyxl

def op_toexcel(data,filename,sheet): # openpyxl库储存数据到excel
    wb = openpyxl.load_workbook(filename) # 创建工作簿对象
    ws = wb.create_sheet(title=sheet,index=0) # 创建子表
    for each in data:
        ws.append(each) # 每次写入一行
    wb.save(filename)



def read_data(file_name,name_index):
    with open(file_name,'r') as file:
        text = file.readlines()
        text = [each.strip() for each in text]
        index = np.array([text.index('blade'),text.index('sumppoint'),text.index('volutepoint')])
        sump_data = []
        volute_data = []
        blade_data = []
        #读取blade中的数据
        for each in text[index[0]+3:index[1]-2]:
            each = each.split(',')
            blade_data.append(each)
        #读取sump中的数据
        for each in text[index[1]+3:index[2]-2]:
            each = each.split(',')
            sump_data.append(each)
        #读取volute中的数据
        for each in text[index[2]+3:]:
            each = each.split(',')
            volute_data.append(each)
    sheet_name = 'time_' + str(name_index)
    op_toexcel(blade_data,r'.\data\blade.xlsx',sheet_name)
    op_toexcel(sump_data,r'.\data\sump.xlsx',sheet_name)
    op_toexcel(volute_data,r'.\data\volute.xlsx',sheet_name)

def delete_sheet(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb["Sheet1"]
    wb.remove(ws)
    wb.save(file_name)

def main():
    file_name_first = r'.\origin_data\export_'
    file_name_last = '.csv'
    for name_index in range(81):
        file_name = file_name_first + str(name_index) + file_name_last
        read_data(file_name,name_index)
    delete_sheet(r'.\data\blade.xlsx')
    delete_sheet(r'.\data\sump.xlsx')
    delete_sheet(r'.\data\volute.xlsx')



if __name__ == '__main__':
    main()